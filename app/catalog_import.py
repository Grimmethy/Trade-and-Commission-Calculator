import importlib
import json
import re
import sqlite3
import sys
from pathlib import Path

import openpyxl

PRICE_RE = re.compile(r"[\d.]+")

EXPECTED_HEADER = ("#", "Item Name", "Price", "Website Link", "Image", "Image URL")

IP_DISPLAY_NAMES = {
    "40k": "Warhammer 40,000",
    "aos": "Age of Sigmar",
    "hh": "Horus Heresy",
}


def parse_price(raw) -> float | None:
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    match = PRICE_RE.search(str(raw).replace(",", ""))
    if not match:
        return None
    try:
        return float(match.group())
    except ValueError:
        return None


def load_faction_lookup(scraper_root: Path) -> dict[str, tuple[str, str]]:
    """Returns {xlsx basename: (ip_display_name, faction_title)}, sourced from
    WarhammerScraper's own factions/*.py registry — the real source of truth for
    which army a product sheet belongs to and its display title, since the data/
    folder layout on disk doesn't reliably reflect either (e.g. an AoS "Chaos"
    folder holds several unrelated armies, and a top-level data/Destruction/
    duplicates a file that actually lives under data/WH Sigmar/Destruction/)."""
    scraper_root = str(scraper_root.resolve())
    inserted = scraper_root not in sys.path
    if inserted:
        sys.path.insert(0, scraper_root)
    try:
        factions_pkg = importlib.import_module("factions")
        importlib.reload(factions_pkg)
        system_map = {
            "40k": factions_pkg.FACTIONS_40K,
            "aos": factions_pkg.FACTIONS_AOS,
            "hh": factions_pkg.FACTIONS_HH,
        }
    finally:
        if inserted:
            sys.path.remove(scraper_root)

    lookup: dict[str, tuple[str, str]] = {}
    for system, factions_dict in system_map.items():
        ip_name = IP_DISPLAY_NAMES[system]
        for entry in factions_dict.values():
            basename = Path(entry["output"]).name
            title = entry["title"]
            if "�" in title:  # known mangled-encoding case (Horus Heresy vehicles)
                title = Path(entry["output"]).parent.name
            lookup[basename] = (ip_name, title)
    return lookup


def iter_catalog_rows(source_dir: Path, faction_lookup: dict[str, tuple[str, str]]):
    """Yields dicts for every product row in every recognized .xlsx under source_dir.
    A file is only imported if its basename is a known entry in faction_lookup (built
    from the scraper's own registry) — this both skips non-standard sheets (e.g. a
    personal inventory list with a different column layout) and avoids double-counting
    stray duplicate copies of the same file living in more than one folder."""
    seen_basenames = set()
    for xlsx_path in sorted(source_dir.glob("**/*.xlsx")):
        basename = xlsx_path.name
        if basename not in faction_lookup or basename in seen_basenames:
            continue

        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active

        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header is None or tuple(header[: len(EXPECTED_HEADER)]) != EXPECTED_HEADER:
            wb.close()
            continue  # not a standard build.py catalog sheet — skip it

        seen_basenames.add(basename)
        ip_name, faction = faction_lookup[basename]

        rows = ws.iter_rows(min_row=2, values_only=True)
        for row in rows:
            if not row or not row[1]:
                continue
            # columns: #, Item Name, Price, Website Link, Image, Image URL
            item_name = row[1]
            price_raw = row[2] if len(row) > 2 else None
            website_link = row[3] if len(row) > 3 else None
            image_url = row[5] if len(row) > 5 else None
            models_per_box = row[6] if len(row) > 6 else None

            box_price = parse_price(price_raw)
            if box_price is None:
                continue

            yield {
                "ip": ip_name,
                "faction": faction,
                "item_name": item_name,
                "box_price": box_price,
                "website_link": website_link,
                "image_url": image_url,
                "source_file": basename,
                "models_per_box": int(models_per_box) if models_per_box else None,
            }
        wb.close()


def upsert_rows(conn: sqlite3.Connection, rows: list[dict]) -> int:
    count = 0
    for row in rows:
        conn.execute(
            """
            INSERT INTO catalog_items (ip, faction, item_name, box_price, website_link, image_url, source_file, imported_at)
            VALUES (:ip, :faction, :item_name, :box_price, :website_link, :image_url, :source_file, datetime('now'))
            ON CONFLICT(ip, faction, item_name) DO UPDATE SET
                box_price = excluded.box_price,
                website_link = excluded.website_link,
                image_url = excluded.image_url,
                source_file = excluded.source_file,
                imported_at = datetime('now')
            """,
            row,
        )
        count += 1

        models_per_box = row.get("models_per_box")
        if models_per_box:
            # cursor.lastrowid is unreliable here: SQLite only updates
            # last_insert_rowid() on a real INSERT, not on the ON CONFLICT DO
            # UPDATE path, so it can silently hold a stale rowid from an
            # unrelated earlier row in this loop. Always look the id up explicitly.
            catalog_item_id = conn.execute(
                "SELECT id FROM catalog_items WHERE ip = ? AND faction = ? AND item_name = ?",
                (row["ip"], row["faction"], row["item_name"]),
            ).fetchone()[0]
            conn.execute(
                """
                INSERT INTO catalog_overlay (catalog_item_id, models_per_box, notes, updated_at)
                VALUES (?, ?, 'scraped from warhammer.com', datetime('now'))
                ON CONFLICT(catalog_item_id) DO UPDATE SET
                    models_per_box = excluded.models_per_box,
                    updated_at = datetime('now')
                WHERE catalog_overlay.notes = 'scraped from warhammer.com'
                """,
                (catalog_item_id, models_per_box),
            )
    conn.commit()
    return count


def dump_seed_json(conn: sqlite3.Connection, out_path: Path) -> int:
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        """
        SELECT ci.ip, ci.faction, ci.item_name, ci.box_price, ci.website_link,
               ci.image_url, ci.source_file, co.models_per_box
        FROM catalog_items ci
        LEFT JOIN catalog_overlay co ON co.catalog_item_id = ci.id
        """
    ).fetchall()
    data = [dict(r) for r in rows]
    out_path.write_text(json.dumps(data, indent=2), encoding="utf-8")
    return len(data)


def load_seed_json(conn: sqlite3.Connection, seed_path: Path) -> int:
    data = json.loads(seed_path.read_text(encoding="utf-8"))
    return upsert_rows(conn, data)
